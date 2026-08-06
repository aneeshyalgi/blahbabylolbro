import openpyxl
from typing import Dict, List, Any, Set, Optional
import re
from datetime import datetime
from dataclasses import dataclass
from models import TableRegion, ColumnInfo, ColumnDataType


@dataclass
class DetectionConfig:
    """Configuration for table detection matching TypeScript version"""
    min_table_size: Dict[str, int]
    max_gap_tolerance: float
    header_confidence_threshold: float
    data_consistency_threshold: float
    min_separation_gap: int
    enable_sub_table_detection: bool
    
    def __init__(self):
        self.min_table_size = {"rows": 2, "cols": 1}
        self.max_gap_tolerance = 0.2
        self.header_confidence_threshold = 0.6
        self.data_consistency_threshold = 0.7
        self.min_separation_gap = 2
        self.enable_sub_table_detection = True


class TableDetector:
    """Advanced table detector matching TypeScript functionality exactly"""
    
    def __init__(self):
        self.config = DetectionConfig()
        self.data_only_workbook = None  # For accessing calculated values
    
    def detect_tables(self, workbook: openpyxl.Workbook, data_only_workbook: Optional[openpyxl.Workbook] = None) -> List[TableRegion]:
        """Main table detection method matching TypeScript version"""
        self.data_only_workbook = data_only_workbook  # Store for use in column analysis
        tables = []
        table_id = 0
        
        # First detect native tables (Excel Table objects and named ranges)
        native_tables = self._detect_native_tables(workbook)
        tables.extend(native_tables)
        table_id += len(native_tables)
        
        # Then detect data-based tables in each sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_tables = self._detect_tables_in_sheet(worksheet, sheet_name, table_id)
            tables.extend(sheet_tables)
            table_id += len(sheet_tables)
        
        # Split tables by gaps if enabled
        if self.config.enable_sub_table_detection:
            tables = self._split_tables_by_gaps(tables, workbook)
        
        # Merge overlapping tables
        return self._merge_overlapping_tables(tables)
    
    def _detect_native_tables(self, workbook: openpyxl.Workbook) -> List[TableRegion]:
        """Detect native Excel tables and named ranges"""
        tables = []
        table_id = 0
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Check for named ranges that might be tables
            if hasattr(workbook, 'defined_names') and workbook.defined_names:
                for name in workbook.defined_names:
                    if name.destinations:
                        for title, coord in name.destinations:
                            if title == sheet_name:
                                try:
                                    # Parse coordinate range
                                    cell_range = worksheet[coord]
                                    if hasattr(cell_range, 'bounds'):
                                        min_col, min_row, max_col, max_row = cell_range.bounds
                                        
                                        region = {
                                            "start_row": min_row - 1,  # Convert to 0-based
                                            "end_row": max_row - 1,
                                            "start_col": min_col - 1,
                                            "end_col": max_col - 1
                                        }
                                        
                                        header_row_count = self._detect_header_row_count(worksheet, region)
                                        columns = self._analyze_table_columns(worksheet, region, header_row_count)
                                        
                                        tables.append(TableRegion(
                                            id=f"table_{table_id}",
                                            name=name.name or "Named Range Table",
                                            sheet=sheet_name,
                                            start_row=region["start_row"],
                                            end_row=region["end_row"],
                                            start_col=region["start_col"],
                                            end_col=region["end_col"],
                                            columns=columns,
                                            confidence=1.0  # Native tables get maximum confidence
                                        ))
                                        table_id += 1
                                except Exception:
                                    # Invalid range, skip
                                    continue
        
        return tables
    
    def _detect_tables_in_sheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                               sheet_name: str, start_id: int) -> List[TableRegion]:
        """Detect tables in a single sheet using advanced algorithms"""
        tables = []
        
        if worksheet.max_row <= 1 or worksheet.max_column <= 1:
            return tables
        
        processed_cells = set()
        
        # Scan sheet for table candidates
        for row in range(worksheet.max_row - self.config.min_table_size["rows"] + 1):
            for col in range(worksheet.max_column - self.config.min_table_size["cols"] + 1):
                cell_key = f"{row},{col}"
                
                if cell_key in processed_cells:
                    continue
                
                # Check if this cell could be start of a table
                cell = worksheet.cell(row=row+1, column=col+1)  # openpyxl is 1-indexed
                if cell.value is not None and str(cell.value).strip():
                    # Expand table from this seed
                    table_candidate = self._expand_table_from_seed(worksheet, row, col)
                    
                    if self._is_valid_table(table_candidate, worksheet):
                        header_row_count = self._detect_header_row_count(worksheet, table_candidate)
                        columns = self._analyze_table_columns(worksheet, table_candidate, header_row_count)
                        style_confidence = self._assess_styled_region_confidence(worksheet, table_candidate)
                        formula_confidence = self._assess_formula_based_confidence(worksheet, table_candidate)
                        
                        confidence = self._calculate_table_confidence(
                            table_candidate, columns, worksheet, style_confidence, formula_confidence
                        )
                        
                        if confidence >= self.config.header_confidence_threshold:
                            table = TableRegion(
                                id=f"table_{start_id + len(tables)}",
                                name=self._generate_table_name(table_candidate, columns),
                                sheet=sheet_name,
                                start_row=table_candidate["start_row"],
                                end_row=table_candidate["end_row"], 
                                start_col=table_candidate["start_col"],
                                end_col=table_candidate["end_col"],
                                columns=columns,
                                confidence=confidence
                            )
                            
                            tables.append(table)
                            
                            # Mark cells as processed
                            for r in range(table_candidate["start_row"], table_candidate["end_row"] + 1):
                                for c in range(table_candidate["start_col"], table_candidate["end_col"] + 1):
                                    processed_cells.add(f"{r},{c}")
        
        return tables
    
    def _expand_table_from_seed(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                               seed_row: int, seed_col: int) -> Dict[str, int]:
        """Expand table boundaries from seed cell matching TypeScript logic"""
        start_row = seed_row
        start_col = seed_col
        end_row = seed_row  
        end_col = seed_col
        
        # Expand columns (right) with gap tolerance
        consecutive_empty_cols = 0
        for col in range(seed_col, worksheet.max_column):
            has_data_in_column = self._column_has_data(
                worksheet, col, seed_row, min(seed_row + 10, worksheet.max_row - 1)
            )
            
            if has_data_in_column:
                end_col = col
                consecutive_empty_cols = 0
            else:
                consecutive_empty_cols += 1
                if consecutive_empty_cols >= 2:
                    break
        
        # Expand rows (down) with gap tolerance
        consecutive_empty_rows = 0
        for row in range(seed_row, worksheet.max_row):
            has_data_in_row = self._row_has_data(worksheet, row, start_col, end_col)
            
            if has_data_in_row:
                end_row = row
                consecutive_empty_rows = 0
            else:
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= 3:
                    break
        
        return {
            "start_row": start_row,
            "end_row": end_row,
            "start_col": start_col, 
            "end_col": end_col
        }
    
    def _column_has_data(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                        col: int, start_row: int, end_row: int) -> bool:
        """Check if column has any data in given row range"""
        for row in range(start_row, end_row + 1):
            cell = worksheet.cell(row=row+1, column=col+1)  # openpyxl is 1-indexed
            if cell.value is not None and str(cell.value).strip():
                return True
        return False
    
    def _row_has_data(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                     row: int, start_col: int, end_col: int) -> bool:
        """Check if row has any data in given column range"""
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row+1, column=col+1)  # openpyxl is 1-indexed
            if cell.value is not None and str(cell.value).strip():
                return True
        return False
    
    def _is_valid_table(self, candidate: Dict[str, int], 
                       worksheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        """Validate table candidate matches TypeScript logic"""
        rows = candidate["end_row"] - candidate["start_row"] + 1
        cols = candidate["end_col"] - candidate["start_col"] + 1
        
        if rows < self.config.min_table_size["rows"] or cols < self.config.min_table_size["cols"]:
            return False
        
        # Check data density
        total_cells = rows * cols
        filled_cells = 0
        
        for row in range(candidate["start_row"], candidate["end_row"] + 1):
            for col in range(candidate["start_col"], candidate["end_col"] + 1):
                cell = worksheet.cell(row=row+1, column=col+1)
                if cell.value is not None and str(cell.value).strip():
                    filled_cells += 1
        
        density = filled_cells / total_cells
        return density >= (1 - self.config.max_gap_tolerance)
    
    def _detect_header_row_count(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                region: Dict[str, int]) -> int:
        """Detect header row count matching TypeScript logic exactly"""
        header_rows = 1
        max_header_rows = min(3, region["end_row"] - region["start_row"])
        
        for header_candidate in range(2, max_header_rows + 1):
            candidate_row = region["start_row"] + header_candidate - 1
            header_like_count = 0
            data_like_count = 0
            
            for col in range(region["start_col"], region["end_col"] + 1):
                cell = worksheet.cell(row=candidate_row+1, column=col+1)  # openpyxl is 1-indexed
                
                if cell.value is not None:
                    value = str(cell.value)
                    
                    # Header-like: string, length < 30, has letters, and is bold
                    is_header_like = (
                        isinstance(cell.value, str) and
                        len(value) < 30 and
                        re.search(r'[a-zA-Z]', value) and
                        (cell.font and cell.font.bold if cell.font else False)
                    )
                    
                    # Data-like: number, date, or long string
                    is_data_like = (
                        isinstance(cell.value, (int, float)) or
                        isinstance(cell.value, datetime) or
                        (isinstance(cell.value, str) and len(value) > 30)
                    )
                    
                    if is_header_like:
                        header_like_count += 1
                    if is_data_like:
                        data_like_count += 1
            
            total_cols = region["end_col"] - region["start_col"] + 1
            
            if header_like_count / total_cols > 0.5 and data_like_count / total_cols < 0.3:
                header_rows = header_candidate
            else:
                break
        
        return header_rows
    
    def _analyze_table_columns(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                              region: Dict[str, int], header_row_count: int = 1) -> List[ColumnInfo]:
        """Analyze table columns matching TypeScript implementation exactly"""
        columns = []
        
        for col in range(region["start_col"], region["end_col"] + 1):
            # Build column name from header rows
            header_parts = []
            
            for header_row in range(header_row_count):
                header_cell = worksheet.cell(
                    row=region["start_row"] + header_row + 1, 
                    column=col + 1
                )  # openpyxl is 1-indexed
                
                if header_cell.value is not None:
                    header_value = str(header_cell.value).strip()
                    if header_value and header_value not in header_parts:
                        header_parts.append(header_value)
            
            # Create column name
            column_name = " - ".join(header_parts) if header_parts else f"Column {col - region['start_col'] + 1}"
            column_name = self._clean_column_name(column_name)
            
            # Collect data values - prioritize calculated values over formulas
            values = []
            null_count = 0
            data_start_row = region["start_row"] + header_row_count
            total_data_rows = region["end_row"] - data_start_row + 1
            
            for row in range(data_start_row, region["end_row"] + 1):
                cell = worksheet.cell(row=row+1, column=col+1)  # openpyxl is 1-indexed
                cell_value = None
                
                # Check if cell has any content at all (including formulas)
                if cell.value is not None or cell.data_type == 'f':
                    # First, try to get calculated value from data-only workbook for formulas
                    if self.data_only_workbook and cell.data_type == 'f':
                        try:
                            data_only_worksheet = self.data_only_workbook[worksheet.title]
                            data_only_cell = data_only_worksheet.cell(row=row+1, column=col+1)
                            if data_only_cell.value is not None and data_only_cell.value != "":
                                cell_value = data_only_cell.value
                        except:
                            pass  # Fallback to regular cell handling
                    
                    # If we still don't have a value, use the original cell value
                    if cell_value is None:
                        if cell.data_type == 'f':  # Formula cell
                            # Try to get cached calculated value
                            if hasattr(cell, 'value') and cell.value is not None:
                                # If it's a calculated result, use it
                                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                    cell_value = cell.value
                                else:
                                    # It's a formula string - for completeness, treat as non-null
                                    # but use a placeholder for type inference
                                    cell_value = "FORMULA_PLACEHOLDER"
                            else:
                                # Formula with no cached value - still count as having data
                                cell_value = "FORMULA_PLACEHOLDER"
                        else:
                            # Regular cell
                            cell_value = cell.value
                    
                    # Add to values if it's meaningful
                    if cell_value is not None:
                        str_value = str(cell_value).strip()
                        # Skip empty strings and error values, but include calculated results
                        if str_value and not str_value.startswith('#'):
                            # Don't include placeholder values in type analysis
                            if cell_value != "FORMULA_PLACEHOLDER":
                                values.append(cell_value)
                        else:
                            null_count += 1
                    else:
                        null_count += 1
                else:
                    null_count += 1
            
            # Infer data type
            data_type = self._infer_column_data_type(values)
            
            # Filter sample values to exclude placeholders and formula strings
            display_values = [
                v for v in values 
                if v != "FORMULA_PLACEHOLDER" and not (isinstance(v, str) and v.startswith('='))
            ]
            sample_values = display_values[:min(5, len(display_values))]
            
            column_info = ColumnInfo(
                name=column_name,
                index=col,  # Keep original column index like TypeScript
                data_type=data_type,
                sample_values=sample_values,
                null_count=null_count,
                total_count=total_data_rows
            )
            
            columns.append(column_info)
        
        return columns
    
    def _clean_column_name(self, name: str) -> str:
        """Clean column names matching TypeScript logic exactly"""
        if not name:
            return "Unnamed Column"
        
        # Remove common prefixes/suffixes
        name = re.sub(r'^(column|col|field|attr)\s*', '', name, flags=re.IGNORECASE)
        name = re.sub(r'\s*(column|col|field|attr)$', '', name, flags=re.IGNORECASE)
        
        # Clean whitespace
        name = re.sub(r'\s+', ' ', name.strip())
        
        # Truncate if too long
        if len(name) > 50:
            name = name[:50]
        
        return name or "Unnamed Column"
    
    def _infer_column_data_type(self, values: List[Any]) -> ColumnDataType:
        """Infer column data type matching TypeScript logic exactly"""
        if not values:
            return ColumnDataType.STRING
        
        type_counts = {
            "number": 0,
            "string": 0,
            "boolean": 0,
            "date": 0
        }
        
        for value in values:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                type_counts["number"] += 1
            elif isinstance(value, bool):
                type_counts["boolean"] += 1
            elif isinstance(value, datetime) or self._is_date_string(value):
                type_counts["date"] += 1
            else:
                type_counts["string"] += 1
        
        total = len(values)
        threshold = 0.8  # 80% threshold like TypeScript
        
        # Check each type against threshold in same order as TypeScript
        if type_counts["number"] / total >= threshold:
            return ColumnDataType.NUMBER
        if type_counts["date"] / total >= threshold:
            return ColumnDataType.DATE
        if type_counts["boolean"] / total >= threshold:
            return ColumnDataType.BOOLEAN
        if type_counts["string"] / total >= threshold:
            return ColumnDataType.STRING
        
        return ColumnDataType.MIXED
    
    def _is_date_string(self, value: Any) -> bool:
        """Check if string represents a date matching TypeScript patterns"""
        if not isinstance(value, str):
            return False
        
        date_patterns = [
            r'^\d{4}-\d{2}-\d{2}$',           # YYYY-MM-DD
            r'^\d{2}/\d{2}/\d{4}$',           # MM/DD/YYYY
            r'^\d{2}-\d{2}-\d{4}$',           # MM-DD-YYYY
            r'^\d{1,2}/\d{1,2}/\d{2,4}$'     # M/D/YY or MM/DD/YYYY
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, value):
                try:
                    # Try to parse the date
                    import datetime as dt
                    dt.datetime.strptime(value.replace('/', '-'), '%m-%d-%Y')
                    return True
                except:
                    try:
                        dt.datetime.strptime(value, '%Y-%m-%d')
                        return True
                    except:
                        continue
        
        return False
    
    def _assess_styled_region_confidence(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                        region: Dict[str, int]) -> float:
        """Assess styling confidence matching TypeScript logic"""
        style_score = 0.0
        cells_checked = 0
        bordered_cells = 0
        colored_cells = 0
        bold_header_cells = 0
        
        # Check header row for bold formatting
        for col in range(region["start_col"], region["end_col"] + 1):
            cell = worksheet.cell(row=region["start_row"]+1, column=col+1)
            if cell.font and cell.font.bold:
                bold_header_cells += 1
        
        # Sample cells for border and color analysis
        sample_rows = [
            region["start_row"],
            (region["start_row"] + region["end_row"]) // 2,
            region["end_row"]
        ]
        sample_cols = [
            region["start_col"],
            (region["start_col"] + region["end_col"]) // 2, 
            region["end_col"]
        ]
        
        for row_idx in sample_rows:
            for col_idx in sample_cols:
                if row_idx > region["end_row"] or col_idx > region["end_col"]:
                    continue
                
                cell = worksheet.cell(row=row_idx+1, column=col_idx+1)
                cells_checked += 1
                
                # Check for borders
                if cell.border:
                    has_border = any([
                        cell.border.top and cell.border.top.style,
                        cell.border.bottom and cell.border.bottom.style,
                        cell.border.left and cell.border.left.style,
                        cell.border.right and cell.border.right.style
                    ])
                    if has_border:
                        bordered_cells += 1
                
                # Check for background colors
                if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                    if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                        colored_cells += 1
        
        # Calculate style scores matching TypeScript
        header_cols = region["end_col"] - region["start_col"] + 1
        
        if bold_header_cells / header_cols > 0.5:
            style_score += 0.3
        elif bold_header_cells / header_cols > 0.25:
            style_score += 0.15
        
        if cells_checked > 0:
            border_ratio = bordered_cells / cells_checked
            if border_ratio > 0.6:
                style_score += 0.25
            elif border_ratio > 0.3:
                style_score += 0.15
        
        if cells_checked > 0:
            color_ratio = colored_cells / cells_checked
            if color_ratio > 0.3:
                style_score += 0.15
            elif color_ratio > 0.1:
                style_score += 0.08
        
        return style_score
    
    def _assess_formula_based_confidence(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                        region: Dict[str, int]) -> float:
        """Assess formula confidence matching TypeScript logic"""
        formula_score = 0.0
        total_cells = 0
        cells_with_formulas = 0
        formula_columns = set()
        
        # Check each column for formulas (skip header row)
        for col in range(region["start_col"], region["end_col"] + 1):
            column_has_formulas = False
            
            for row in range(region["start_row"] + 1, region["end_row"] + 1):  # Skip header
                cell = worksheet.cell(row=row+1, column=col+1)
                total_cells += 1
                
                if cell.data_type == 'f':  # Formula cell
                    cells_with_formulas += 1
                    column_has_formulas = True
            
            if column_has_formulas:
                formula_columns.add(col)
        
        # Score based on formula presence matching TypeScript
        if cells_with_formulas > 0:
            formula_ratio = cells_with_formulas / total_cells if total_cells > 0 else 0
            
            # Tables often have entire columns with formulas
            if len(formula_columns) > 0:
                column_ratio = len(formula_columns) / (region["end_col"] - region["start_col"] + 1)
                if column_ratio >= 0.3:
                    formula_score += 0.2  # 30% or more columns have formulas
                elif column_ratio >= 0.1:
                    formula_score += 0.1  # 10-30% columns have formulas
            
            # Score based on overall formula density
            if formula_ratio >= 0.3:
                formula_score += 0.15  # High formula density
            elif formula_ratio >= 0.1:
                formula_score += 0.1   # Medium formula density  
            elif formula_ratio > 0:
                formula_score += 0.05  # Some formulas present
        
        return formula_score
    
    def _calculate_table_confidence(self, region: Dict[str, int], columns: List[ColumnInfo],
                                   worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                   style_confidence: float, formula_confidence: float) -> float:
        """Calculate table confidence matching TypeScript logic exactly"""
        confidence = 0.5  # Base confidence
        
        # Size bonus
        cell_count = (region["end_row"] - region["start_row"] + 1) * (region["end_col"] - region["start_col"] + 1)
        if cell_count >= 20:
            confidence += 0.15
        if cell_count >= 50:
            confidence += 0.1
        
        # Header quality
        header_quality = self._assess_header_quality(worksheet, region, columns)
        confidence += header_quality * 0.2
        
        # Data consistency
        data_consistency = self._assess_data_consistency(columns)
        confidence += data_consistency * 0.15
        
        # Type diversity
        unique_types = len(set(col.data_type for col in columns))
        if unique_types >= 2:
            confidence += 0.1
        
        # Add style and formula confidence
        confidence += style_confidence + formula_confidence
        
        return min(confidence, 1.0)
    
    def _assess_header_quality(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                              region: Dict[str, int], columns: List[ColumnInfo]) -> float:
        """Assess header quality matching TypeScript logic"""
        score = 0.0
        header_row = region["start_row"]
        
        for col in range(region["start_col"], region["end_col"] + 1):
            cell = worksheet.cell(row=header_row+1, column=col+1)
            
            if cell.value is not None:
                header_value = str(cell.value)
                
                # String headers are good
                if isinstance(cell.value, str):
                    score += 0.3
                
                # Reasonable length
                if 2 <= len(header_value) <= 30:
                    score += 0.2
                
                # Contains letters
                if re.search(r'[a-zA-Z]', header_value):
                    score += 0.3
                
                # Not purely numeric
                if not re.match(r'^\d+$', header_value):
                    score += 0.2
        
        return score / len(columns) if columns else 0.0
    
    def _assess_data_consistency(self, columns: List[ColumnInfo]) -> float:
        """Assess data consistency matching TypeScript logic"""
        consistency_score = 0.0
        
        for column in columns:
            # Non-mixed types are good
            if column.data_type != ColumnDataType.MIXED:
                consistency_score += 0.5
            
            # Low null rate is good
            null_rate = column.null_count / column.total_count if column.total_count > 0 else 0
            if null_rate < 0.1:
                consistency_score += 0.3
            elif null_rate < 0.3:
                consistency_score += 0.2
            
            # Having sample values is good
            if column.sample_values:
                consistency_score += 0.2
        
        return consistency_score / len(columns) if columns else 0.0
    
    def _generate_table_name(self, region: Dict[str, int], columns: List[ColumnInfo]) -> str:
        """Generate table name matching TypeScript logic"""
        meaningful_headers = [
            col.name for col in columns
            if len(col.name) > 1 and not re.match(r'^Column \d+$', col.name)
        ][:2]  # Take first 2 meaningful headers
        
        if meaningful_headers:
            return " & ".join(meaningful_headers) + " Table"
        
        # Fallback to position-based name
        start_cell_col = chr(ord('A') + region["start_col"]) if region["start_col"] < 26 else f"Col{region['start_col']}"
        start_cell = f"{start_cell_col}{region['start_row'] + 1}"
        return f"Table at {start_cell}"
    
    def _split_tables_by_gaps(self, tables: List[TableRegion], workbook: openpyxl.Workbook) -> List[TableRegion]:
        """Split tables by gaps matching TypeScript logic"""
        split_tables = []
        
        for table in tables:
            worksheet = workbook[table.sheet]
            sub_tables = self._find_sub_tables_in_region(worksheet, table)
            
            if len(sub_tables) > 1:
                split_tables.extend(sub_tables)
            else:
                split_tables.append(table)
        
        return split_tables
    
    def _find_sub_tables_in_region(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                  region: TableRegion) -> List[TableRegion]:
        """Find sub-tables within region matching TypeScript logic"""
        # First split horizontally by empty rows
        horizontal_splits = self._split_by_empty_rows(worksheet, region)
        
        # Then split each horizontally by empty columns
        all_sub_tables = []
        for h_split in horizontal_splits:
            vertical_splits = self._split_by_empty_columns(worksheet, h_split)
            all_sub_tables.extend(vertical_splits)
        
        # Filter out tables that are too small
        return [
            table for table in all_sub_tables
            if ((table.end_row - table.start_row + 1) >= self.config.min_table_size["rows"] and
                (table.end_col - table.start_col + 1) >= self.config.min_table_size["cols"])
        ]
    
    def _split_by_empty_rows(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                            region: TableRegion) -> List[TableRegion]:
        """Split by empty rows matching TypeScript logic"""
        empty_rows = []
        
        # Find completely empty rows
        for row in range(region.start_row, region.end_row + 1):
            if not self._row_has_data(worksheet, row, region.start_col, region.end_col):
                empty_rows.append(row)
        
        # Find gaps of consecutive empty rows
        gaps = []
        gap_start = -1
        
        for i, row in enumerate(empty_rows):
            if gap_start == -1:
                gap_start = row
            
            if i == len(empty_rows) - 1 or empty_rows[i + 1] != row + 1:
                gap_end = row
                gap_size = gap_end - gap_start + 1
                
                if gap_size >= self.config.min_separation_gap:
                    gaps.append({"start": gap_start, "end": gap_end})
                gap_start = -1
        
        # If no significant gaps, return original
        if not gaps:
            return [region]
        
        # Split by gaps
        sub_regions = []
        current_start = region.start_row
        
        for gap in gaps:
            if gap["start"] > current_start:
                sub_region = TableRegion(
                    id=f"{region.id}_split_{len(sub_regions)}",
                    name=f"{region.name} (Part {len(sub_regions) + 1})",
                    sheet=region.sheet,
                    start_row=current_start,
                    end_row=gap["start"] - 1,
                    start_col=region.start_col,
                    end_col=region.end_col,
                    columns=region.columns,
                    confidence=region.confidence
                )
                sub_regions.append(sub_region)
            current_start = gap["end"] + 1
        
        # Add final region
        if current_start <= region.end_row:
            sub_region = TableRegion(
                id=f"{region.id}_split_{len(sub_regions)}",
                name=f"{region.name} (Part {len(sub_regions) + 1})",
                sheet=region.sheet,
                start_row=current_start,
                end_row=region.end_row,
                start_col=region.start_col,
                end_col=region.end_col,
                columns=region.columns,
                confidence=region.confidence
            )
            sub_regions.append(sub_region)
        
        return sub_regions
    
    def _split_by_empty_columns(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                               region: TableRegion) -> List[TableRegion]:
        """Split by empty columns matching TypeScript logic"""
        empty_cols = []
        
        # Find completely empty columns
        for col in range(region.start_col, region.end_col + 1):
            if not self._column_has_data(worksheet, col, region.start_row, region.end_row):
                empty_cols.append(col)
        
        # Find gaps of consecutive empty columns
        gaps = []
        gap_start = -1
        
        for i, col in enumerate(empty_cols):
            if gap_start == -1:
                gap_start = col
            
            if i == len(empty_cols) - 1 or empty_cols[i + 1] != col + 1:
                gap_end = col
                gap_size = gap_end - gap_start + 1
                
                if gap_size >= self.config.min_separation_gap:
                    gaps.append({"start": gap_start, "end": gap_end})
                gap_start = -1
        
        # If no significant gaps, return original
        if not gaps:
            return [region]
        
        # Split by gaps
        sub_regions = []
        current_start = region.start_col
        
        for gap in gaps:
            if gap["start"] > current_start:
                sub_region = TableRegion(
                    id=f"{region.id}_colsplit_{len(sub_regions)}",
                    name=f"{region.name} (Col Part {len(sub_regions) + 1})",
                    sheet=region.sheet,
                    start_row=region.start_row,
                    end_row=region.end_row,
                    start_col=current_start,
                    end_col=gap["start"] - 1,
                    columns=region.columns,
                    confidence=region.confidence
                )
                sub_regions.append(sub_region)
            current_start = gap["end"] + 1
        
        # Add final region
        if current_start <= region.end_col:
            sub_region = TableRegion(
                id=f"{region.id}_colsplit_{len(sub_regions)}",
                name=f"{region.name} (Col Part {len(sub_regions) + 1})",
                sheet=region.sheet,
                start_row=region.start_row,
                end_row=region.end_row,
                start_col=current_start,
                end_col=region.end_col,
                columns=region.columns,
                confidence=region.confidence
            )
            sub_regions.append(sub_region)
        
        return sub_regions
    
    def _merge_overlapping_tables(self, tables: List[TableRegion]) -> List[TableRegion]:
        """Merge overlapping tables matching TypeScript logic"""
        sorted_tables = sorted(tables, key=lambda t: t.confidence, reverse=True)
        merged_tables = []
        
        for table in sorted_tables:
            merged = False
            
            for i, existing in enumerate(merged_tables):
                if self._tables_overlap(table, existing) and table.sheet == existing.sheet:
                    if not self._are_tables_separated(table, existing):
                        # Tables overlap and aren't separated - keep higher confidence one
                        if table.confidence > existing.confidence:
                            merged_tables[i] = table
                        merged = True
                        break
            
            if not merged:
                merged_tables.append(table)
        
        return merged_tables
    
    def _tables_overlap(self, table1: TableRegion, table2: TableRegion) -> bool:
        """Check if tables overlap matching TypeScript logic"""
        overlap = self._calculate_overlap(table1, table2)
        table1_size = (table1.end_row - table1.start_row + 1) * (table1.end_col - table1.start_col + 1)
        table2_size = (table2.end_row - table2.start_row + 1) * (table2.end_col - table2.start_col + 1)
        
        overlap_ratio = overlap / min(table1_size, table2_size) if min(table1_size, table2_size) > 0 else 0
        return overlap_ratio > 0.5
    
    def _calculate_overlap(self, table1: TableRegion, table2: TableRegion) -> int:
        """Calculate overlap area matching TypeScript logic"""
        row_overlap = max(0, min(table1.end_row, table2.end_row) - max(table1.start_row, table2.start_row) + 1)
        col_overlap = max(0, min(table1.end_col, table2.end_col) - max(table1.start_col, table2.start_col) + 1)
        
        return row_overlap * col_overlap
    
    def _are_tables_separated(self, table1: TableRegion, table2: TableRegion) -> bool:
        """Check if tables are separated by gaps matching TypeScript logic"""
        # Check vertical separation
        if table1.end_row < table2.start_row:
            gap = table2.start_row - table1.end_row - 1
            return gap >= self.config.min_separation_gap
        
        if table2.end_row < table1.start_row:
            gap = table1.start_row - table2.end_row - 1
            return gap >= self.config.min_separation_gap
        
        # Check horizontal separation
        if table1.end_col < table2.start_col:
            gap = table2.start_col - table1.end_col - 1
            return gap >= self.config.min_separation_gap
        
        if table2.end_col < table1.start_col:
            gap = table1.start_col - table2.end_col - 1
            return gap >= self.config.min_separation_gap
        
        return False
